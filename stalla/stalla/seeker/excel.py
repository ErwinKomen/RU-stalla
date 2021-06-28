"""Reading and writing of Excel files for the Stalla app

Reading - this is so as to download (or update) data
Writing - this is to export the current data for the researchers
"""
from stalla.utils import ErrHandle
from stalla.settings import WRITABLE_DIR, MEDIA_DIR

import io, sys, os
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.cell import Cell
from openpyxl import Workbook
from io import StringIO

def excel_to_list(filename, wsName = None, lExpected = None, lField = None):
    """Read an excel file into a list of objects

    This assumes that the first row contains column headers
    The [wsName] may contain the name of the worksheet to be read
    """

    oErr = ErrHandle()
    bResult = True
    lData = []
    msg = ""
    try:
        # Read file
        wb = openpyxl.load_workbook(filename, read_only=True)
        if wsName == None or wsName == "":
            ws = wb.active
        else:
            # Try to open the indicated worksheet
            sheetnames = wb.sheetnames
            if wsName in sheetnames:
                ws = wb[wsName]
            else:
                # we are not able to read this one
                bResult = False
                msg = "The sheet [{}] doesn't exist".format(wsName)
                return bResult, lData, msg

        # Iterate through rows
        bFirst = True
        
        lHeader = []
        # Iterate
        iRow = 1
        for row in ws.iter_rows(min_row=1, min_col=1):
            oErr.Status("Reading row {}".format(iRow))
            iRow += 1
            if bFirst:
                # Expect header
                for cell in row:
                    sValue = cell.value.strip("\t").lower()                    
                    if lExpected == None or lField == None:
                        sKey = sValue
                    else:
                        sKey = ""
                        for idx, item in enumerate(lExpected):
                            if item in sValue:
                                sKey = lField[idx]
                                break
                        # Check if it's okay
                        if sKey == "":
                            # Cannot read this
                            msg = "Don't understand column header [{}]".format(sValue)
                            return False, [], msg
                    lHeader.append(sKey)
                bFirst = False
            elif row[0].value != None:
                oRow = {}
                for idx, key in enumerate(lHeader):
                    cell = row[idx]
                    # Get the value as a string
                    cv = "" if cell.value == None else "{}".format(cell.value).strip()
                    oRow[key] = cv
                # Also add the row number (as string)
                oRow['row_number'] = "{}".format(row[0].row)
                lData.append(oRow)
            else:
                break

        # Close the workbook
        wb.close()

        ## Remove the file
        #os.remove(tmp_path)
        # Return positively
        bResult = True
    except:
        # Note the error here
        msg = oErr.get_error_message()
        bResult = False
        oErr.DoError("excel_to_list")

    # Return what we have found
    return bResult, lData, msg


def excel_generator(filename, wsName = None, lExpected = None, lField = None):
    """Read an excel file into a list of objects

    This assumes that the first row contains column headers
    The [wsName] may contain the name of the worksheet to be read
    """

    oErr = ErrHandle()
    bResult = True
    lData = []
    msg = ""
    try:
        # Read file
        wb = openpyxl.load_workbook(filename, read_only=True)
        if wsName == None or wsName == "":
            ws = wb.active
        else:
            # Try to open the indicated worksheet
            sheetnames = wb.sheetnames
            if wsName in sheetnames:
                ws = wb[wsName]
            else:
                # we are not able to read this one
                bResult = False
                msg = "The sheet [{}] doesn't exist".format(wsName)
                return bResult, lData, msg

        # Iterate through rows
        bFirst = True
        
        lHeader = []
        # Iterate
        for row in ws.iter_rows(min_row=1, min_col=1):
            if bFirst:
                # Expect header
                for cell in row:
                    sValue = cell.value.strip("\t").lower()                    
                    if lExpected == None or lField == None:
                        sKey = sValue
                    else:
                        sKey = ""
                        for idx, item in enumerate(lExpected):
                            if item in sValue:
                                sKey = lField[idx]
                                break
                        # Check if it's okay
                        if sKey == "":
                            # Cannot read this
                            msg = "Don't understand column header [{}]".format(sValue)
                            return False, [], msg
                    lHeader.append(sKey)
                bFirst = False
            elif row[0].value != None:
                oRow = {}
                for idx, key in enumerate(lHeader):
                    cell = row[idx]
                    # Get the value as a string
                    cv = "" if cell.value == None else "{}".format(cell.value).strip()
                    oRow[key] = cv
                # Also add the row number (as string)
                oRow['row_number'] = "{}".format(row[0].row)

                # Use the YIELD function to return data
                yield oRow

        # Close the workbook
        wb.close()

        ## Remove the file
        #os.remove(tmp_path)
        # Return positively
        bResult = True
    except:
        # Note the error here
        msg = oErr.get_error_message()
        bResult = False
        oErr.DoError("excel_generator")

    # Return the last part
    return None

